from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import date, timedelta
from typing import Optional
import io
from ..database import get_db
from ..models.sales import SalesOrder
from ..models.purchase import PurchaseOrder
from ..models.manufacturing import ManufacturingOrder
from ..models.product import Product
from ..models.inventory import StockLedger
from ..models.user import User
from ..middleware.auth_middleware import get_current_user

router = APIRouter(prefix="/reports", tags=["Reports"])


@router.get("/sales")
def sales_report(
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not from_date:
        from_date = date.today() - timedelta(days=30)
    if not to_date:
        to_date = date.today()

    orders = db.query(SalesOrder).filter(
        SalesOrder.order_date >= from_date,
        SalesOrder.order_date <= to_date,
    ).order_by(SalesOrder.order_date).all()

    total_value = sum(float(o.total_amount or 0) for o in orders)
    by_status = {}
    for o in orders:
        by_status[o.status] = by_status.get(o.status, 0) + 1

    return {
        "from_date": str(from_date),
        "to_date": str(to_date),
        "total_orders": len(orders),
        "total_value": round(total_value, 2),
        "by_status": by_status,
        "orders": [
            {
                "order_number": o.order_number,
                "order_date": str(o.order_date),
                "customer": o.customer.company_name if o.customer else None,
                "status": o.status,
                "total_amount": float(o.total_amount or 0),
            }
            for o in orders
        ],
    }


@router.get("/purchase")
def purchase_report(
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not from_date:
        from_date = date.today() - timedelta(days=30)
    if not to_date:
        to_date = date.today()

    orders = db.query(PurchaseOrder).filter(
        PurchaseOrder.order_date >= from_date,
        PurchaseOrder.order_date <= to_date,
    ).order_by(PurchaseOrder.order_date).all()

    total_value = sum(float(o.total_amount or 0) for o in orders)
    return {
        "from_date": str(from_date),
        "to_date": str(to_date),
        "total_orders": len(orders),
        "total_value": round(total_value, 2),
        "orders": [
            {
                "order_number": o.order_number,
                "order_date": str(o.order_date),
                "vendor": o.vendor.company_name if o.vendor else None,
                "status": o.status,
                "total_amount": float(o.total_amount or 0),
                "auto_generated": o.auto_generated,
            }
            for o in orders
        ],
    }


@router.get("/inventory")
def inventory_report(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    products = db.query(Product).filter(Product.is_active == True).all()
    return {
        "total_products": len(products),
        "total_value": round(sum(float(p.on_hand_qty or 0) * float(p.cost_price or 0) for p in products), 2),
        "low_stock_count": sum(1 for p in products if p.is_low_stock),
        "products": [
            {
                "product_code": p.product_code,
                "product_name": p.product_name,
                "on_hand_qty": float(p.on_hand_qty or 0),
                "reserved_qty": float(p.reserved_qty or 0),
                "free_qty": p.free_qty,
                "cost_price": float(p.cost_price or 0),
                "inventory_value": float(p.on_hand_qty or 0) * float(p.cost_price or 0),
                "is_low_stock": p.is_low_stock,
            }
            for p in products
        ],
    }


@router.get("/manufacturing")
def manufacturing_report(
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not from_date:
        from_date = date.today() - timedelta(days=30)
    if not to_date:
        to_date = date.today()

    orders = db.query(ManufacturingOrder).filter(
        ManufacturingOrder.created_at >= from_date,
    ).order_by(ManufacturingOrder.created_at.desc()).all()

    by_status = {}
    for o in orders:
        by_status[o.status] = by_status.get(o.status, 0) + 1

    return {
        "from_date": str(from_date),
        "to_date": str(to_date),
        "total_orders": len(orders),
        "by_status": by_status,
        "orders": [
            {
                "mo_number": o.mo_number,
                "product": o.product.product_name if o.product else None,
                "planned_qty": float(o.planned_qty or 0),
                "produced_qty": float(o.produced_qty or 0),
                "status": o.status,
                "auto_generated": o.auto_generated,
            }
            for o in orders
        ],
    }


@router.get("/profit")
def profit_report(
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not from_date:
        from_date = date.today() - timedelta(days=30)
    if not to_date:
        to_date = date.today()

    from ..models.sales import SalesOrderItem
    items = (
        db.query(SalesOrderItem)
        .join(SalesOrder)
        .filter(
            SalesOrder.order_date >= from_date,
            SalesOrder.order_date <= to_date,
            SalesOrder.status != "cancelled",
        )
        .all()
    )

    total_revenue = sum(float(i.line_total or 0) for i in items)
    total_cost = sum(float(i.ordered_qty or 0) * float(i.product.cost_price or 0) for i in items if i.product)
    gross_profit = total_revenue - total_cost
    margin = (gross_profit / total_revenue * 100) if total_revenue else 0

    return {
        "from_date": str(from_date),
        "to_date": str(to_date),
        "total_revenue": round(total_revenue, 2),
        "total_cost": round(total_cost, 2),
        "gross_profit": round(gross_profit, 2),
        "gross_margin_pct": round(margin, 2),
    }


@router.get("/export/sales/excel")
def export_sales_excel(
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not from_date:
        from_date = date.today() - timedelta(days=30)
    if not to_date:
        to_date = date.today()

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    headers = ["Order#", "Date", "Customer", "Status", "Subtotal", "Tax", "Total"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    orders = db.query(SalesOrder).filter(
        SalesOrder.order_date >= from_date,
        SalesOrder.order_date <= to_date,
    ).all()

    for row, o in enumerate(orders, 2):
        ws.cell(row=row, column=1, value=o.order_number)
        ws.cell(row=row, column=2, value=str(o.order_date))
        ws.cell(row=row, column=3, value=o.customer.company_name if o.customer else "")
        ws.cell(row=row, column=4, value=o.status)
        ws.cell(row=row, column=5, value=float(o.subtotal or 0))
        ws.cell(row=row, column=6, value=float(o.tax_amount or 0))
        ws.cell(row=row, column=7, value=float(o.total_amount or 0))

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=sales_report_{from_date}_{to_date}.xlsx"},
    )
