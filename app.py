"""
MINI ERP - Shiv Furniture
Flask Backend + Static HTML/CSS/JS Frontend

Run: python app.py
Open: http://localhost:8000
Login: admin / Admin@123
"""

from flask import Flask, request, jsonify, send_from_directory, redirect
from flask_cors import CORS
import pymysql, pymysql.cursors
import bcrypt
import jwt
import os
import re
from datetime import datetime, timedelta, date
from decimal import Decimal
from functools import wraps

# ── App Config ────────────────────────────────────────────────────────
BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
STATIC_DIR  = os.path.join(BASE_DIR, 'static')

app = Flask(__name__, static_folder=STATIC_DIR)
CORS(app, resources={r"/api/*": {"origins": "*"}})

DB = dict(host='localhost', user='root', password='', database='mini_erp',
          charset='utf8mb4', cursorclass=pymysql.cursors.DictCursor,
          autocommit=False)

SECRET_KEY = 'mini-erp-shiv-furniture-secret-key-2024'

# ── DB Helper ─────────────────────────────────────────────────────────
def get_db():
    return pymysql.connect(**DB)

def qry(sql, args=(), one=False, commit=False):
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute(sql, args)
            if commit:
                conn.commit()
                return cur.lastrowid
            return cur.fetchone() if one else cur.fetchall()
    finally:
        conn.close()

def paginate_list(items, page=1, page_size=15):
    total = len(items)
    start = (page-1)*page_size
    return {
        'items': items[start:start+page_size],
        'meta': {'page': page, 'page_size': page_size, 'total': total, 'pages': max(1,(total+page_size-1)//page_size)}
    }

def serial(obj):
    """Make objects JSON-serialisable."""
    if isinstance(obj, dict):
        return {k: serial(v) for k,v in obj.items()}
    if isinstance(obj, list):
        return [serial(v) for v in obj]
    if isinstance(obj, (datetime, date)):
        return obj.isoformat()
    if isinstance(obj, Decimal):
        return float(obj)
    return obj

def ok(data={}, code=200):
    return jsonify(serial(data)), code

def err(msg, code=400):
    return jsonify({'detail': msg}), code

def next_seq(prefix):
    """Generate next sequence number using sequence_counters table."""
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT id, last_number FROM sequence_counters WHERE prefix=%s FOR UPDATE", (prefix,))
            row = cur.fetchone()
            if not row:
                cur.execute("INSERT INTO sequence_counters(prefix,last_number,updated_at) VALUES(%s,1,NOW())", (prefix,))
                n = 1
            else:
                n = (row['last_number'] or 0) + 1
                cur.execute("UPDATE sequence_counters SET last_number=%s, updated_at=NOW() WHERE prefix=%s", (n, prefix))
            conn.commit()
        return f"{prefix}-{str(n).zfill(5)}"
    except Exception:
        conn.rollback()
        import time
        return f"{prefix}-{int(time.time())}"
    finally:
        conn.close()

# ── Role Permissions Map ──────────────────────────────────────────────
ROLE_WRITE = {
    'admin':              ['all'],
    'business_owner':     [],           # read-only
    'sales_user':         ['sales','customers','deliveries'],
    'purchase_user':      ['purchase','vendors'],
    'manufacturing_user': ['manufacturing','boms'],
    'inventory_manager':  ['inventory','products'],
}

def _user_can_write(user_roles, module):
    """Return True if any of the user's roles allows writing to module."""
    if 'admin' in user_roles:
        return True
    for role in user_roles:
        perms = ROLE_WRITE.get(role, [])
        if 'all' in perms or module in perms:
            return True
    return False

def require_write(module):
    """Decorator: reject if user cannot write to module. Must be placed AFTER @token_required."""
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            roles = getattr(request, 'current_user', {}).get('roles', [])
            if not _user_can_write(roles, module):
                return err(f'Access denied. Your role does not have write access to {module}.', 403)
            return f(*args, **kwargs)
        return decorated
    return decorator

# ── Auth Middleware ───────────────────────────────────────────────────
def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = None
        auth = request.headers.get('Authorization','')
        if auth.startswith('Bearer '):
            token = auth.split(' ',1)[1]
        if not token:
            token = request.args.get('token')
        if not token:
            return err('Unauthorized', 401)
        try:
            data = jwt.decode(token, SECRET_KEY, algorithms=['HS256'])
            uid  = int(data['sub'])
            user = qry("SELECT u.*, GROUP_CONCAT(r.name) as role_names FROM users u LEFT JOIN user_roles ur ON u.id=ur.user_id LEFT JOIN roles r ON ur.role_id=r.id WHERE u.id=%s GROUP BY u.id", (uid,), one=True)
            if not user:
                return err('User not found', 401)
            request.current_user = user
            request.current_user['roles'] = [n for n in (user.get('role_names') or '').split(',') if n]
        except jwt.ExpiredSignatureError:
            return err('Token expired', 401)
        except Exception:
            return err('Invalid token', 401)
        return f(*args, **kwargs)
    return decorated

# ── Static Files ──────────────────────────────────────────────────────
@app.route('/')
def root():
    return redirect('/static/index.html')

@app.route('/static/')
def static_index():
    return send_from_directory(STATIC_DIR, 'index.html')

@app.route('/static/<path:filename>')
def serve_static(filename):
    return send_from_directory(STATIC_DIR, filename)

# ── AUTH ──────────────────────────────────────────────────────────────
@app.route('/api/v1/auth/login', methods=['POST'])
def login():
    data = request.get_json() or {}
    username = data.get('username','').strip()
    password = data.get('password','')
    if not username or not password:
        return err('Username and password required')

    user = qry("""SELECT u.*, GROUP_CONCAT(r.name SEPARATOR ',') as role_names,
                  GROUP_CONCAT(r.id SEPARATOR ',') as role_ids
                  FROM users u
                  LEFT JOIN user_roles ur ON u.id=ur.user_id
                  LEFT JOIN roles r ON ur.role_id=r.id
                  WHERE (u.username=%s OR u.email=%s) AND u.is_active=1
                  GROUP BY u.id""", (username, username), one=True)

    if not user:
        return err('Invalid credentials', 401)

    stored = user['password_hash'] or ''
    try:
        ok_pass = bcrypt.checkpw(password.encode(), stored.encode() if isinstance(stored,str) else stored)
    except Exception:
        ok_pass = False

    if not ok_pass:
        return err('Invalid credentials', 401)

    # Update last login
    qry("UPDATE users SET last_login=NOW() WHERE id=%s", (user['id'],), commit=True)

    exp = datetime.utcnow() + timedelta(hours=24)
    token = jwt.encode({'sub': str(user['id']), 'username': user['username'], 'exp': exp}, SECRET_KEY, algorithm='HS256')
    ref_token = jwt.encode({'sub': str(user['id']), 'exp': datetime.utcnow()+timedelta(days=7)}, SECRET_KEY, algorithm='HS256')

    roles = [{'id': int(i), 'name': n} for i,n in zip(
        (user.get('role_ids') or '').split(','),
        (user.get('role_names') or '').split(',')
    ) if i and n]

    return ok({
        'access_token': token,
        'refresh_token': ref_token,
        'user': {
            'id': user['id'], 'username': user['username'],
            'email': user['email'], 'full_name': user['full_name'],
            'phone': user['phone'], 'is_active': bool(user['is_active']),
            'last_login': user['last_login'], 'created_at': user['created_at'],
            'roles': roles
        }
    })

@app.route('/api/v1/auth/change-password', methods=['POST'])
@token_required
def change_password():
    data = request.get_json() or {}
    cur_pwd = data.get('current_password','')
    new_pwd = data.get('new_password','')
    if not cur_pwd or not new_pwd or len(new_pwd) < 8:
        return err('Invalid password data')
    user = qry("SELECT * FROM users WHERE id=%s", (request.current_user['id'],), one=True)
    if not bcrypt.checkpw(cur_pwd.encode(), (user['password_hash'] or '').encode()):
        return err('Current password incorrect')
    h = bcrypt.hashpw(new_pwd.encode(), bcrypt.gensalt()).decode()
    qry("UPDATE users SET password_hash=%s WHERE id=%s", (h, user['id']), commit=True)
    return ok({'message': 'Password updated'})

# ── DASHBOARD ─────────────────────────────────────────────────────────
@app.route('/api/v1/dashboard/stats')
@token_required
def dashboard_stats():
    total_products  = qry("SELECT COUNT(*) as c FROM products WHERE is_active=1", one=True)['c']
    low_stock       = qry("SELECT COUNT(*) as c FROM products WHERE is_active=1 AND on_hand_qty<=reorder_point AND on_hand_qty>0 AND reorder_point>0", one=True)['c']
    out_of_stock    = qry("SELECT COUNT(*) as c FROM products WHERE is_active=1 AND on_hand_qty<=0", one=True)['c']
    total_customers = qry("SELECT COUNT(*) as c FROM customers WHERE is_active=1", one=True)['c']
    total_vendors   = qry("SELECT COUNT(*) as c FROM vendors WHERE is_active=1", one=True)['c']
    pending_so      = qry("SELECT COUNT(*) as c FROM sales_orders WHERE status IN('draft','confirmed')", one=True)['c']
    pending_po      = qry("SELECT COUNT(*) as c FROM purchase_orders WHERE status IN('draft','confirmed')", one=True)['c']
    active_mo       = qry("SELECT COUNT(*) as c FROM manufacturing_orders WHERE status IN('confirmed','in_progress')", one=True)['c']
    monthly_rev     = qry("SELECT COALESCE(SUM(total_amount),0) as r FROM sales_orders WHERE order_date>=DATE_SUB(NOW(),INTERVAL 30 DAY) AND status!='cancelled'", one=True)['r']
    so_month        = qry("SELECT COUNT(*) as c FROM sales_orders WHERE order_date>=DATE_SUB(NOW(),INTERVAL 30 DAY)", one=True)['c']
    total_so        = qry("SELECT COUNT(*) as c FROM sales_orders", one=True)['c']
    total_po        = qry("SELECT COUNT(*) as c FROM purchase_orders", one=True)['c']

    trend = qry("""SELECT DATE_FORMAT(order_date,'%%b %%Y') as month, SUM(total_amount) as revenue, COUNT(*) as orders
                   FROM sales_orders WHERE order_date>=DATE_SUB(NOW(),INTERVAL 12 MONTH) AND status!='cancelled'
                   GROUP BY DATE_FORMAT(order_date,'%%Y-%%m') ORDER BY MIN(order_date)""")

    inv_cat = qry("""SELECT COALESCE(c.name,'Uncategorized') as category_name,
                     SUM(p.on_hand_qty * p.cost_price) as total_value
                     FROM products p LEFT JOIN categories c ON p.category_id=c.id
                     WHERE p.is_active=1 GROUP BY c.name ORDER BY total_value DESC LIMIT 6""")

    low_prods = qry("""SELECT product_name, product_code, on_hand_qty, reorder_point
                       FROM products WHERE is_active=1 AND on_hand_qty<=reorder_point AND reorder_point>0
                       ORDER BY on_hand_qty LIMIT 8""")

    return ok({
        'kpis': {
            'total_products': total_products, 'low_stock_items': low_stock,
            'out_of_stock': out_of_stock, 'total_customers': total_customers,
            'total_vendors': total_vendors, 'pending_sales_orders': pending_so,
            'pending_purchase_orders': pending_po, 'active_manufacturing_orders': active_mo,
            'monthly_revenue': float(monthly_rev or 0), 'so_this_month': so_month,
            'total_so': total_so, 'total_po': total_po
        },
        'sales_trend': trend,
        'inventory_by_category': inv_cat,
        'low_stock_products': low_prods
    })

@app.route('/api/v1/dashboard/recent-activity')
@token_required
def recent_activity():
    sales = qry("""SELECT so.id, so.order_number, so.order_date, so.total_amount, so.status,
                   c.company_name as customer_name FROM sales_orders so
                   LEFT JOIN customers c ON so.customer_id=c.id
                   ORDER BY so.created_at DESC LIMIT 6""")
    return ok({'sales_orders': sales})

@app.route('/api/v1/dashboard/notifications')
@token_required
def notifications():
    limit = int(request.args.get('limit', 8))
    items = qry("""SELECT id, type, message, is_read, created_at FROM notifications
                   ORDER BY created_at DESC LIMIT %s""", (limit,))
    return ok({'notifications': items})

# ── PRODUCTS ──────────────────────────────────────────────────────────
@app.route('/api/v1/products/', methods=['GET'])
@app.route('/api/v1/products', methods=['GET'])
@token_required
def list_products():
    page      = int(request.args.get('page', 1))
    page_size = int(request.args.get('page_size', 20))
    search    = request.args.get('search', '')
    is_active = request.args.get('is_active', '1')
    proc_type = request.args.get('procurement_type', '')

    where, args = ['1=1'], []
    if is_active != '': where.append('p.is_active=%s'); args.append(int(is_active) if is_active in('0','1') else 1)
    if search: where.append('(p.product_name LIKE %s OR p.product_code LIKE %s)'); args += [f'%{search}%', f'%{search}%']
    if proc_type: where.append('p.procurement_type=%s'); args.append(proc_type)

    sql = f"""SELECT p.*, p.sales_price as sale_price, COALESCE(c.name,'') as category
              FROM products p LEFT JOIN categories c ON p.category_id=c.id
              WHERE {' AND '.join(where)} ORDER BY p.product_name"""
    items = qry(sql, args)
    return ok(paginate_list(items, page, page_size))

@app.route('/api/v1/products/<int:pid>', methods=['GET'])
@token_required
def get_product(pid):
    p = qry("""SELECT p.*, p.sales_price as sale_price, COALESCE(c.name,'') as category
               FROM products p LEFT JOIN categories c ON p.category_id=c.id
               WHERE p.id=%s""", (pid,), one=True)
    return ok(p) if p else err('Not found', 404)

def get_or_create_category(name):
    """Return category_id for given name, creating if needed."""
    if not name: return None
    cat = qry("SELECT id FROM categories WHERE name=%s", (name,), one=True)
    if cat: return cat['id']
    return qry("INSERT INTO categories(name) VALUES(%s)", (name,), commit=True)

@app.route('/api/v1/products/', methods=['POST'])
@app.route('/api/v1/products', methods=['POST'])
@token_required
@require_write('products')
def create_product():
    d = request.get_json() or {}
    if not d.get('product_name'): return err('Product name required')
    cat_id = get_or_create_category(d.get('category',''))
    rid = qry("""INSERT INTO products(product_code,product_name,category_id,unit_of_measure,
                 sales_price,cost_price,reorder_point,description,procurement_type,is_active,created_by)
                 VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,1,%s)""",
              (d.get('product_code',''), d['product_name'], cat_id,
               d.get('unit_of_measure','PCS'), d.get('sale_price',0),
               d.get('cost_price',0), d.get('reorder_point',0),
               d.get('description',''), d.get('procurement_type','buy'),
               request.current_user['id']), commit=True)
    p = qry("""SELECT p.*, p.sales_price as sale_price, COALESCE(c.name,'') as category
               FROM products p LEFT JOIN categories c ON p.category_id=c.id WHERE p.id=%s""", (rid,), one=True)
    return ok(p, 201)

@app.route('/api/v1/products/<int:pid>', methods=['PUT'])
@token_required
@require_write('products')
def update_product(pid):
    d = request.get_json() or {}
    cat_id = get_or_create_category(d.get('category',''))
    qry("""UPDATE products SET product_code=%s,product_name=%s,category_id=%s,unit_of_measure=%s,
           sales_price=%s,cost_price=%s,reorder_point=%s,description=%s,procurement_type=%s WHERE id=%s""",
        (d.get('product_code',''), d.get('product_name',''), cat_id,
         d.get('unit_of_measure','PCS'), d.get('sale_price',0), d.get('cost_price',0),
         d.get('reorder_point',0), d.get('description',''), d.get('procurement_type','buy'), pid), commit=True)
    return ok(qry("""SELECT p.*, p.sales_price as sale_price, COALESCE(c.name,'') as category
                     FROM products p LEFT JOIN categories c ON p.category_id=c.id WHERE p.id=%s""", (pid,), one=True))

@app.route('/api/v1/products/<int:pid>', methods=['DELETE'])
@token_required
@require_write('products')
def delete_product(pid):
    qry("UPDATE products SET is_active=0 WHERE id=%s", (pid,), commit=True)
    return '', 204

@app.route('/api/v1/products/<int:pid>/adjust-stock', methods=['POST'])
@token_required
@require_write('inventory')
def adjust_stock(pid):
    d    = request.get_json() or {}
    qty  = float(d.get('quantity', 0))
    note = d.get('notes', 'Manual adjustment')
    p    = qry("SELECT * FROM products WHERE id=%s", (pid,), one=True)
    if not p: return err('Product not found', 404)

    new_qty = float(p.get('on_hand_qty') or 0) + qty
    if new_qty < 0: return err('Insufficient stock')

    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("UPDATE products SET on_hand_qty=%s WHERE id=%s", (new_qty, pid))
            cur.execute("""INSERT INTO stock_ledger(product_id,movement_date,movement_type,quantity,
                          balance_qty,unit_cost,reference_type,reference_id,reference_number,notes,created_by)
                          VALUES(%s,NOW(),'manual_adjustment',%s,%s,%s,'manual',%s,'ADJ',%s,%s)""",
                       (pid, qty, new_qty, p.get('cost_price',0), pid, note, request.current_user['id']))
            conn.commit()
        return ok({'message': 'Stock adjusted', 'new_quantity': new_qty})
    except Exception as e:
        conn.rollback(); return err(str(e))
    finally:
        conn.close()

# ── CUSTOMERS ─────────────────────────────────────────────────────────
@app.route('/api/v1/sales/customers', methods=['GET'])
@token_required
def list_customers():
    page = int(request.args.get('page',1)); ps = int(request.args.get('page_size',20))
    s = request.args.get('search','')
    where, args = ['is_active=1'], []
    if s: where.append('(company_name LIKE %s OR contact_name LIKE %s OR email LIKE %s)'); args += [f'%{s}%']*3
    items = qry(f"SELECT * FROM customers WHERE {' AND '.join(where)} ORDER BY company_name", args)
    return ok(paginate_list(items, page, ps))

@app.route('/api/v1/sales/customers/<int:cid>', methods=['GET'])
@token_required
def get_customer(cid):
    c = qry("SELECT * FROM customers WHERE id=%s",(cid,),one=True)
    return ok(c) if c else err('Not found',404)

@app.route('/api/v1/sales/customers', methods=['POST'])
@token_required
@require_write('customers')
def create_customer():
    d = request.get_json() or {}
    if not d.get('company_name'): return err('Company name required')
    code = f"CUST-{qry('SELECT COUNT(*)+1 as n FROM customers',one=True)['n']:04d}"
    rid = qry("""INSERT INTO customers(customer_code,company_name,contact_name,email,phone,address,city,state,country,gst_number,is_active)
                 VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,1)""",
              (code,d['company_name'],d.get('contact_name',''),d.get('email',''),d.get('phone',''),
               d.get('address',''),d.get('city',''),d.get('state',''),d.get('country','India'),d.get('gst_number','')), commit=True)
    return ok(qry("SELECT * FROM customers WHERE id=%s",(rid,),one=True), 201)

@app.route('/api/v1/sales/customers/<int:cid>', methods=['PUT'])
@token_required
@require_write('customers')
def update_customer(cid):
    d = request.get_json() or {}
    qry("UPDATE customers SET company_name=%s,contact_name=%s,email=%s,phone=%s,address=%s,city=%s,gst_number=%s WHERE id=%s",
        (d.get('company_name',''),d.get('contact_name',''),d.get('email',''),d.get('phone',''),
         d.get('address',''),d.get('city',''),d.get('gst_number',''),cid), commit=True)
    return ok(qry("SELECT * FROM customers WHERE id=%s",(cid,),one=True))

@app.route('/api/v1/sales/customers/<int:cid>', methods=['DELETE'])
@token_required
@require_write('customers')
def delete_customer(cid):
    qry("UPDATE customers SET is_active=0 WHERE id=%s",(cid,),commit=True)
    return '', 204

# ── VENDORS ───────────────────────────────────────────────────────────
@app.route('/api/v1/purchase/vendors', methods=['GET'])
@token_required
def list_vendors():
    page=int(request.args.get('page',1)); ps=int(request.args.get('page_size',20))
    s=request.args.get('search','')
    where,args=['is_active=1'],[]
    if s: where.append('(company_name LIKE %s OR contact_name LIKE %s)'); args+=[f'%{s}%']*2
    items=qry(f"SELECT * FROM vendors WHERE {' AND '.join(where)} ORDER BY company_name",args)
    return ok(paginate_list(items,page,ps))

@app.route('/api/v1/purchase/vendors/<int:vid>', methods=['GET'])
@token_required
def get_vendor(vid):
    v=qry("SELECT * FROM vendors WHERE id=%s",(vid,),one=True)
    return ok(v) if v else err('Not found',404)

@app.route('/api/v1/purchase/vendors', methods=['POST'])
@token_required
@require_write('vendors')
def create_vendor():
    d=request.get_json() or {}
    if not d.get('company_name'): return err('Company name required')
    code=f"VEND-{qry('SELECT COUNT(*)+1 as n FROM vendors',one=True)['n']:04d}"
    rid=qry("""INSERT INTO vendors(vendor_code,company_name,contact_name,email,phone,address,city,state,country,gst_number,lead_time_days,is_active)
               VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,1)""",
            (code,d['company_name'],d.get('contact_name',''),d.get('email',''),d.get('phone',''),
             d.get('address',''),d.get('city',''),d.get('state',''),d.get('country','India'),
             d.get('gst_number',''),d.get('lead_time_days',7)),commit=True)
    return ok(qry("SELECT * FROM vendors WHERE id=%s",(rid,),one=True),201)

@app.route('/api/v1/purchase/vendors/<int:vid>', methods=['PUT'])
@token_required
@require_write('vendors')
def update_vendor(vid):
    d=request.get_json() or {}
    qry("UPDATE vendors SET company_name=%s,contact_name=%s,email=%s,phone=%s,address=%s,city=%s,gst_number=%s,lead_time_days=%s WHERE id=%s",
        (d.get('company_name',''),d.get('contact_name',''),d.get('email',''),d.get('phone',''),
         d.get('address',''),d.get('city',''),d.get('gst_number',''),d.get('lead_time_days',7),vid),commit=True)
    return ok(qry("SELECT * FROM vendors WHERE id=%s",(vid,),one=True))

@app.route('/api/v1/purchase/vendors/<int:vid>', methods=['DELETE'])
@token_required
@require_write('vendors')
def delete_vendor(vid):
    qry("UPDATE vendors SET is_active=0 WHERE id=%s",(vid,),commit=True)
    return '',204

# ── SALES ORDERS ──────────────────────────────────────────────────────
@app.route('/api/v1/sales/orders', methods=['GET'])
@token_required
def list_sales_orders():
    page=int(request.args.get('page',1)); ps=int(request.args.get('page_size',15))
    s=request.args.get('search',''); status=request.args.get('status','')
    where,args=[],[]
    if s: where.append('(so.order_number LIKE %s OR c.company_name LIKE %s)'); args+=[f'%{s}%']*2
    if status: where.append('so.status=%s'); args.append(status)
    w = 'WHERE '+' AND '.join(where) if where else ''
    items=qry(f"""SELECT so.*, c.company_name as customer_name,
                  JSON_OBJECT('id',c.id,'company_name',c.company_name) as customer
                  FROM sales_orders so LEFT JOIN customers c ON so.customer_id=c.id
                  {w} ORDER BY so.created_at DESC""", args)
    for item in items:
        if isinstance(item.get('customer'), str):
            import json
            try: item['customer'] = json.loads(item['customer'])
            except: pass
    return ok(paginate_list(items, page, ps))

@app.route('/api/v1/sales/orders', methods=['POST'])
@token_required
@require_write('sales')
def create_sales_order():
    d = request.get_json() or {}
    if not d.get('customer_id'): return err('Customer required')
    if not d.get('items'): return err('Items required')

    conn = get_db()
    try:
        with conn.cursor() as cur:
            order_num = next_seq('SO')
            cur.execute("""INSERT INTO sales_orders(order_number,customer_id,order_date,expected_delivery_date,
                          notes,status,subtotal,tax_amount,discount_amount,total_amount,created_by)
                          VALUES(%s,%s,%s,%s,%s,'draft',0,0,0,0,%s)""",
                       (order_num, d['customer_id'], d.get('order_date', date.today().isoformat()),
                        d.get('expected_delivery_date'), d.get('notes',''), request.current_user['id']))
            so_id = cur.lastrowid
            total = 0
            for item in d['items']:
                pid   = item['product_id']
                qty   = float(item.get('ordered_qty',1))
                price = float(item.get('unit_price',0))
                tax   = float(item.get('tax_pct',0))
                disc  = float(item.get('discount_pct',0))
                base  = qty*price*(1-disc/100)
                ltotal= base*(1+tax/100)
                total += ltotal
                cur.execute("""INSERT INTO sales_order_items(sales_order_id,product_id,ordered_qty,
                              delivered_qty,unit_price,discount_pct,tax_pct,line_total)
                              VALUES(%s,%s,%s,0,%s,%s,%s,%s)""",
                           (so_id,pid,qty,price,disc,tax,ltotal))
            cur.execute("UPDATE sales_orders SET subtotal=%s,total_amount=%s WHERE id=%s",(total,total,so_id))
            conn.commit()
        so = qry("""SELECT so.*, c.company_name as customer_name FROM sales_orders so
                    LEFT JOIN customers c ON so.customer_id=c.id WHERE so.id=%s""", (so_id,), one=True)
        return ok(so, 201)
    except Exception as e:
        conn.rollback(); return err(str(e))
    finally:
        conn.close()

@app.route('/api/v1/sales/orders/<int:oid>/confirm', methods=['POST'])
@token_required
@require_write('sales')
def confirm_sales_order(oid):
    so = qry("SELECT * FROM sales_orders WHERE id=%s",(oid,),one=True)
    if not so: return err('Not found',404)
    if so['status'] != 'draft': return err(f"Cannot confirm order in status: {so['status']}")
    qry("UPDATE sales_orders SET status='confirmed', confirmed_at=NOW() WHERE id=%s",(oid,),commit=True)
    return ok({'message': 'Order confirmed'})

@app.route('/api/v1/sales/orders/<int:oid>/cancel', methods=['POST'])
@token_required
@require_write('sales')
def cancel_sales_order(oid):
    so = qry("SELECT * FROM sales_orders WHERE id=%s",(oid,),one=True)
    if not so: return err('Not found',404)
    if so['status'] in ('delivered','cancelled'): return err(f"Cannot cancel order in status: {so['status']}")
    qry("UPDATE sales_orders SET status='cancelled' WHERE id=%s",(oid,),commit=True)
    return ok({'message': 'Order cancelled'})

# ── PURCHASE ORDERS ───────────────────────────────────────────────────
@app.route('/api/v1/purchase/orders', methods=['GET'])
@token_required
def list_purchase_orders():
    page=int(request.args.get('page',1)); ps=int(request.args.get('page_size',15))
    s=request.args.get('search',''); status=request.args.get('status','')
    where,args=[],[]
    if s: where.append('(po.order_number LIKE %s OR v.company_name LIKE %s)'); args+=[f'%{s}%']*2
    if status: where.append('po.status=%s'); args.append(status)
    w='WHERE '+' AND '.join(where) if where else ''
    items=qry(f"""SELECT po.*, v.company_name as vendor_name,
                  JSON_OBJECT('id',v.id,'company_name',v.company_name) as vendor
                  FROM purchase_orders po LEFT JOIN vendors v ON po.vendor_id=v.id
                  {w} ORDER BY po.created_at DESC""", args)
    for item in items:
        if isinstance(item.get('vendor'), str):
            import json
            try: item['vendor'] = json.loads(item['vendor'])
            except: pass
    return ok(paginate_list(items,page,ps))

@app.route('/api/v1/purchase/orders', methods=['POST'])
@token_required
@require_write('purchase')
def create_purchase_order():
    d=request.get_json() or {}
    if not d.get('vendor_id'): return err('Vendor required')
    if not d.get('items'): return err('Items required')
    conn=get_db()
    try:
        with conn.cursor() as cur:
            order_num=next_seq('PO')
            cur.execute("""INSERT INTO purchase_orders(order_number,vendor_id,order_date,expected_receipt_date,
                          notes,status,subtotal,tax_amount,total_amount,auto_generated,created_by)
                          VALUES(%s,%s,%s,%s,%s,'draft',0,0,0,0,%s)""",
                       (order_num,d['vendor_id'],d.get('order_date',date.today().isoformat()),
                        d.get('expected_receipt_date'),d.get('notes',''),request.current_user['id']))
            po_id=cur.lastrowid; subtotal=0; tax_total=0
            for item in d['items']:
                qty=float(item.get('ordered_qty',1)); price=float(item.get('unit_price',0))
                tax=float(item.get('tax_pct',0)); ltotal=qty*price; ltax=ltotal*tax/100
                subtotal+=ltotal; tax_total+=ltax
                cur.execute("""INSERT INTO purchase_order_items(purchase_order_id,product_id,ordered_qty,
                              received_qty,unit_price,tax_pct,line_total)
                              VALUES(%s,%s,%s,0,%s,%s,%s)""",
                           (po_id,item['product_id'],qty,price,tax,ltotal+ltax))
            cur.execute("UPDATE purchase_orders SET subtotal=%s,tax_amount=%s,total_amount=%s WHERE id=%s",
                       (subtotal,tax_total,subtotal+tax_total,po_id))
            conn.commit()
        po=qry("SELECT * FROM purchase_orders WHERE id=%s",(po_id,),one=True)
        return ok(po,201)
    except Exception as e:
        conn.rollback(); return err(str(e))
    finally:
        conn.close()

@app.route('/api/v1/purchase/orders/<int:oid>/confirm', methods=['POST'])
@token_required
@require_write('purchase')
def confirm_po(oid):
    po=qry("SELECT * FROM purchase_orders WHERE id=%s",(oid,),one=True)
    if not po: return err('Not found',404)
    if po['status']!='draft': return err(f"Cannot confirm PO in status: {po['status']}")
    qry("UPDATE purchase_orders SET status='confirmed', confirmed_at=NOW() WHERE id=%s",(oid,),commit=True)
    return ok({'message':'PO confirmed'})

@app.route('/api/v1/purchase/orders/<int:oid>/cancel', methods=['POST'])
@token_required
@require_write('purchase')
def cancel_po(oid):
    po=qry("SELECT * FROM purchase_orders WHERE id=%s",(oid,),one=True)
    if not po: return err('Not found',404)
    if po['status'] in('fully_received','cancelled'): return err('Cannot cancel')
    qry("UPDATE purchase_orders SET status='cancelled' WHERE id=%s",(oid,),commit=True)
    return ok({'message':'PO cancelled'})

@app.route('/api/v1/purchase/orders/<int:oid>/receive', methods=['POST'])
@token_required
@require_write('purchase')
def receive_po(oid):
    po=qry("SELECT * FROM purchase_orders WHERE id=%s",(oid,),one=True)
    if not po: return err('Not found',404)
    if po['status']!='confirmed': return err('PO must be confirmed first')
    items=qry("SELECT * FROM purchase_order_items WHERE purchase_order_id=%s",(oid,))
    conn=get_db()
    try:
        with conn.cursor() as cur:
            for item in items:
                pid=item['product_id']; qty=float(item['ordered_qty'])
                p=qry("SELECT * FROM products WHERE id=%s",(pid,),one=True)
                new_qty=float(p.get('on_hand_qty') or 0)+qty
                cur.execute("UPDATE products SET on_hand_qty=%s WHERE id=%s",(new_qty,pid))
                cur.execute("""INSERT INTO stock_ledger(product_id,movement_date,movement_type,quantity,
                              balance_qty,unit_cost,reference_type,reference_id,reference_number,notes,created_by)
                              VALUES(%s,NOW(),'purchase_receipt',%s,%s,%s,'purchase_order',%s,%s,'Goods receipt',%s)""",
                           (pid,qty,new_qty,item.get('unit_price',0),oid,po['order_number'],request.current_user['id']))
                cur.execute("UPDATE purchase_order_items SET received_qty=%s WHERE id=%s",(qty,item['id']))
            cur.execute("UPDATE purchase_orders SET status='fully_received', received_at=NOW() WHERE id=%s",(oid,))
            conn.commit()
        return ok({'message':'Goods received, stock updated'})
    except Exception as e:
        conn.rollback(); return err(str(e))
    finally:
        conn.close()

# ── INVENTORY ─────────────────────────────────────────────────────────
@app.route('/api/v1/inventory/stock-summary')
@token_required
def stock_summary():
    items=qry("""SELECT p.id, p.product_code, p.product_name, p.on_hand_qty, p.reserved_qty,
                 p.reorder_point, p.unit_of_measure, p.cost_price, p.sales_price as sale_price,
                 GREATEST(0, p.on_hand_qty - p.reserved_qty) as free_qty,
                 (p.on_hand_qty <= p.reorder_point) as is_low_stock,
                 (p.on_hand_qty * p.cost_price) as inventory_value,
                 COALESCE(c.name,'') as category
                 FROM products p LEFT JOIN categories c ON p.category_id=c.id
                 WHERE p.is_active=1 ORDER BY p.product_name""")
    return jsonify(serial(items))

@app.route('/api/v1/inventory/ledger')
@token_required
def stock_ledger():
    page=int(request.args.get('page',1)); ps=int(request.args.get('page_size',25))
    mt=request.args.get('movement_type','')
    where,args=['1=1'],[]
    if mt: where.append('sl.movement_type=%s'); args.append(mt)
    items=qry(f"""SELECT sl.*,p.product_name,p.product_code FROM stock_ledger sl
                  LEFT JOIN products p ON sl.product_id=p.id
                  WHERE {' AND '.join(where)} ORDER BY sl.movement_date DESC, sl.id DESC""",args)
    return ok(paginate_list(items,page,ps))

# ── MANUFACTURING ORDERS ──────────────────────────────────────────────
@app.route('/api/v1/manufacturing/orders', methods=['GET'])
@token_required
def list_mo():
    page=int(request.args.get('page',1)); ps=int(request.args.get('page_size',15))
    status=request.args.get('status','')
    where,args=['1=1'],[]
    if status: where.append('status=%s'); args.append(status)
    items=qry(f"SELECT * FROM manufacturing_orders WHERE {' AND '.join(where)} ORDER BY created_at DESC",args)
    return ok(paginate_list(items,page,ps))

@app.route('/api/v1/manufacturing/orders', methods=['POST'])
@token_required
@require_write('manufacturing')
def create_mo():
    d=request.get_json() or {}
    if not d.get('product_id'): return err('Product required')
    mo_num=next_seq('MO')
    rid=qry("""INSERT INTO manufacturing_orders(mo_number,product_id,bom_id,planned_qty,produced_qty,
               scheduled_start,scheduled_end,status,auto_generated,notes,created_by)
               VALUES(%s,%s,%s,%s,0,%s,%s,'draft',0,%s,%s)""",
            (mo_num,d['product_id'],d.get('bom_id'),d.get('planned_qty',1),
             d.get('scheduled_start'),d.get('scheduled_end'),d.get('notes',''),
             request.current_user['id']),commit=True)
    return ok(qry("SELECT * FROM manufacturing_orders WHERE id=%s",(rid,),one=True),201)

@app.route('/api/v1/manufacturing/orders/<int:mid>/confirm', methods=['POST'])
@token_required
@require_write('manufacturing')
def confirm_mo(mid):
    mo=qry("SELECT * FROM manufacturing_orders WHERE id=%s",(mid,),one=True)
    if not mo or mo['status']!='draft': return err('Cannot confirm')
    qry("UPDATE manufacturing_orders SET status='confirmed' WHERE id=%s",(mid,),commit=True)
    return ok({'message':'MO confirmed'})

@app.route('/api/v1/manufacturing/orders/<int:mid>/start', methods=['POST'])
@token_required
@require_write('manufacturing')
def start_mo(mid):
    mo=qry("SELECT * FROM manufacturing_orders WHERE id=%s",(mid,),one=True)
    if not mo or mo['status']!='confirmed': return err('MO must be confirmed')
    qry("UPDATE manufacturing_orders SET status='in_progress',actual_start=NOW() WHERE id=%s",(mid,),commit=True)
    return ok({'message':'Production started'})

@app.route('/api/v1/manufacturing/orders/<int:mid>/done', methods=['POST'])
@token_required
@require_write('manufacturing')
def done_mo(mid):
    mo=qry("SELECT * FROM manufacturing_orders WHERE id=%s",(mid,),one=True)
    if not mo or mo['status']!='in_progress': return err('MO must be in progress')
    qty=float(mo.get('planned_qty') or 0); pid=mo['product_id']
    p=qry("SELECT * FROM products WHERE id=%s",(pid,),one=True)
    if not p: return err('Product not found',404)
    conn=get_db()
    try:
        with conn.cursor() as cur:
            new_qty=float(p.get('on_hand_qty') or 0)+qty
            cur.execute("UPDATE products SET on_hand_qty=%s WHERE id=%s",(new_qty,pid))
            cur.execute("""INSERT INTO stock_ledger(product_id,movement_date,movement_type,quantity,
                          balance_qty,unit_cost,reference_type,reference_id,reference_number,notes,created_by)
                          VALUES(%s,NOW(),'production_output',%s,%s,%s,'manufacturing_order',%s,%s,'MFG done',%s)""",
                       (pid,qty,new_qty,p.get('cost_price',0),mid,mo['mo_number'],request.current_user['id']))
            cur.execute("UPDATE manufacturing_orders SET status='done',produced_qty=%s,actual_end=NOW() WHERE id=%s",(qty,mid))
            conn.commit()
        return ok({'message':'Production done, stock updated'})
    except Exception as e:
        conn.rollback(); return err(str(e))
    finally:
        conn.close()

# ── BOMs ──────────────────────────────────────────────────────────────
@app.route('/api/v1/manufacturing/boms', methods=['GET'])
@token_required
def list_boms():
    page=int(request.args.get('page',1)); ps=int(request.args.get('page_size',15))
    items=qry("SELECT * FROM boms WHERE is_active=1 ORDER BY bom_code")
    for b in items:
        b['components']=qry("SELECT * FROM bom_components WHERE bom_id=%s",(b['id'],))
    return ok(paginate_list(items,page,ps))

@app.route('/api/v1/manufacturing/boms/<int:bid>', methods=['GET'])
@token_required
def get_bom(bid):
    b=qry("SELECT * FROM boms WHERE id=%s",(bid,),one=True)
    if not b: return err('Not found',404)
    b['components']=qry("SELECT * FROM bom_components WHERE bom_id=%s",(bid,))
    return ok(b)

@app.route('/api/v1/manufacturing/boms', methods=['POST'])
@token_required
@require_write('boms')
def create_bom():
    d=request.get_json() or {}
    if not d.get('product_id'): return err('Product required')
    bom_code=next_seq('BOM')
    conn=get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("""INSERT INTO boms(bom_code,product_id,version,quantity,notes,is_active,created_by)
                          VALUES(%s,%s,%s,%s,%s,1,%s)""",
                       (bom_code,d['product_id'],d.get('version','1.0'),d.get('quantity',1),
                        d.get('notes',''),request.current_user['id']))
            bom_id=cur.lastrowid
            for comp in d.get('components',[]):
                cur.execute("""INSERT INTO bom_components(bom_id,component_product_id,quantity,unit_of_measure)
                              VALUES(%s,%s,%s,%s)""",
                           (bom_id,comp['component_product_id'],comp.get('quantity',1),comp.get('unit_of_measure','PCS')))
            conn.commit()
        b=qry("SELECT * FROM boms WHERE id=%s",(bom_id,),one=True)
        b['components']=qry("SELECT * FROM bom_components WHERE bom_id=%s",(bom_id,))
        return ok(b,201)
    except Exception as e:
        conn.rollback(); return err(str(e))
    finally:
        conn.close()

@app.route('/api/v1/manufacturing/boms/<int:bid>', methods=['PUT'])
@token_required
@require_write('boms')
def update_bom(bid):
    d=request.get_json() or {}
    conn=get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("UPDATE boms SET version=%s,quantity=%s,notes=%s WHERE id=%s",
                       (d.get('version','1.0'),d.get('quantity',1),d.get('notes',''),bid))
            cur.execute("DELETE FROM bom_components WHERE bom_id=%s",(bid,))
            for comp in d.get('components',[]):
                cur.execute("INSERT INTO bom_components(bom_id,component_product_id,quantity,unit_of_measure) VALUES(%s,%s,%s,%s)",
                           (bid,comp['component_product_id'],comp.get('quantity',1),comp.get('unit_of_measure','PCS')))
            conn.commit()
        b=qry("SELECT * FROM boms WHERE id=%s",(bid,),one=True)
        b['components']=qry("SELECT * FROM bom_components WHERE bom_id=%s",(bid,))
        return ok(b)
    except Exception as e:
        conn.rollback(); return err(str(e))
    finally:
        conn.close()

@app.route('/api/v1/manufacturing/boms/<int:bid>', methods=['DELETE'])
@token_required
@require_write('boms')
def delete_bom(bid):
    qry("UPDATE boms SET is_active=0 WHERE id=%s",(bid,),commit=True)
    return '',204

# ── USERS ─────────────────────────────────────────────────────────────
@app.route('/api/v1/users/', methods=['GET'])
@app.route('/api/v1/users', methods=['GET'])
@token_required
def list_users():
    page=int(request.args.get('page',1)); ps=int(request.args.get('page_size',15))
    s=request.args.get('search','')
    where,args=['1=1'],[]
    if s: where.append('(u.username LIKE %s OR u.email LIKE %s OR u.full_name LIKE %s)'); args+=[f'%{s}%']*3
    users=qry(f"""SELECT u.id,u.username,u.email,u.full_name,u.phone,u.is_active,u.last_login,u.created_at
                  FROM users u WHERE {' AND '.join(where)} ORDER BY u.username""",args)
    for u in users:
        roles=qry("SELECT r.id,r.name FROM roles r JOIN user_roles ur ON r.id=ur.role_id WHERE ur.user_id=%s",(u['id'],))
        u['roles']=roles
    return ok(paginate_list(users,page,ps))

@app.route('/api/v1/users/<int:uid>', methods=['GET'])
@token_required
def get_user(uid):
    u=qry("SELECT id,username,email,full_name,phone,is_active,last_login,created_at FROM users WHERE id=%s",(uid,),one=True)
    if not u: return err('Not found',404)
    u['roles']=qry("SELECT r.id,r.name FROM roles r JOIN user_roles ur ON r.id=ur.role_id WHERE ur.user_id=%s",(uid,))
    return ok(u)

@app.route('/api/v1/users/', methods=['POST'])
@app.route('/api/v1/users', methods=['POST'])
@token_required
@require_write('users')
def create_user():
    d=request.get_json() or {}
    if not d.get('username') or not d.get('email') or not d.get('password'):
        return err('Username, email and password required')
    h=bcrypt.hashpw(d['password'].encode(),bcrypt.gensalt()).decode()
    conn=get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("""INSERT INTO users(username,email,full_name,phone,password_hash,is_active)
                          VALUES(%s,%s,%s,%s,%s,%s)""",
                       (d['username'],d['email'],d.get('full_name',''),d.get('phone',''),h,d.get('is_active',True)))
            uid=cur.lastrowid
            for rname in d.get('roles',['viewer']):
                role=qry("SELECT id FROM roles WHERE name=%s",(rname,),one=True)
                if role:
                    cur.execute("INSERT IGNORE INTO user_roles(user_id,role_id) VALUES(%s,%s)",(uid,role['id']))
            conn.commit()
        u=qry("SELECT id,username,email,full_name,phone,is_active FROM users WHERE id=%s",(uid,),one=True)
        u['roles']=qry("SELECT r.id,r.name FROM roles r JOIN user_roles ur ON r.id=ur.role_id WHERE ur.user_id=%s",(uid,))
        return ok(u,201)
    except pymysql.err.IntegrityError:
        conn.rollback(); return err('Username or email already exists')
    except Exception as e:
        conn.rollback(); return err(str(e))
    finally:
        conn.close()

@app.route('/api/v1/users/<int:uid>', methods=['PUT'])
@token_required
@require_write('users')
def update_user(uid):
    d=request.get_json() or {}
    conn=get_db()
    try:
        with conn.cursor() as cur:
            if d.get('password'):
                h=bcrypt.hashpw(d['password'].encode(),bcrypt.gensalt()).decode()
                cur.execute("UPDATE users SET username=%s,email=%s,full_name=%s,phone=%s,is_active=%s,password_hash=%s WHERE id=%s",
                           (d.get('username',''),d.get('email',''),d.get('full_name',''),d.get('phone',''),d.get('is_active',True),h,uid))
            else:
                cur.execute("UPDATE users SET username=%s,email=%s,full_name=%s,phone=%s,is_active=%s WHERE id=%s",
                           (d.get('username',''),d.get('email',''),d.get('full_name',''),d.get('phone',''),d.get('is_active',True),uid))
            if 'roles' in d:
                cur.execute("DELETE FROM user_roles WHERE user_id=%s",(uid,))
                for rname in d['roles']:
                    role=qry("SELECT id FROM roles WHERE name=%s",(rname,),one=True)
                    if role: cur.execute("INSERT IGNORE INTO user_roles(user_id,role_id) VALUES(%s,%s)",(uid,role['id']))
            conn.commit()
        conn.commit()
        u = qry("SELECT id,username,email,full_name,phone,is_active,last_login FROM users WHERE id=%s",(uid,),one=True)
        u['roles'] = qry("SELECT r.id,r.name FROM roles r JOIN user_roles ur ON r.id=ur.role_id WHERE ur.user_id=%s",(uid,))
        return ok(u)
    except Exception as e:
        conn.rollback(); return err(str(e))
    finally:
        conn.close()

@app.route('/api/v1/users/<int:uid>', methods=['PATCH'])
@token_required
@require_write('users')
def patch_user(uid):
    d=request.get_json() or {}
    if 'is_active' in d:
        qry("UPDATE users SET is_active=%s WHERE id=%s",(d['is_active'],uid),commit=True)
    return ok({'message':'Updated'})

# ── REPORTS ───────────────────────────────────────────────────────────
@app.route('/api/v1/reports/summary')
@token_required
def reports_summary():
    return dashboard_stats()

@app.route('/api/v1/reports/<rtype>/export')
@token_required
def export_report(rtype):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io
    except ImportError:
        return err('openpyxl not found. Run: pip install openpyxl', 501)

    wb = openpyxl.Workbook()
    ws = wb.active

    # Styles
    hdr_font  = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill  = PatternFill('solid', fgColor='6366F1')
    hdr_align = Alignment(horizontal='center', vertical='center')
    thin      = Side(style='thin', color='E2E8F0')
    bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill  = PatternFill('solid', fgColor='F8FAFC')
    title_font= Font(bold=True, size=14, color='0F172A')
    sub_font  = Font(size=10, color='64748B')

    def style_header(row_num, ncols):
        for c in range(1, ncols+1):
            cell = ws.cell(row=row_num, column=c)
            cell.font    = hdr_font
            cell.fill    = hdr_fill
            cell.alignment = hdr_align
            cell.border  = bdr

    def style_data_row(row_num, ncols, is_alt=False):
        for c in range(1, ncols+1):
            cell = ws.cell(row=row_num, column=c)
            if is_alt: cell.fill = alt_fill
            cell.border  = bdr
            cell.alignment = Alignment(vertical='center')

    def add_title(title, subtitle):
        ws.merge_cells(f'A1:{get_column_letter(8)}1')
        t = ws['A1']; t.value = title; t.font = title_font
        t.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 28
        ws.merge_cells(f'A2:{get_column_letter(8)}2')
        s = ws['A2']; s.value = subtitle; s.font = sub_font
        ws.row_dimensions[2].height = 18

    from datetime import datetime as dt
    now = dt.now().strftime('%d %b %Y %H:%M')

    if rtype == 'sales':
        ws.title = 'Sales Report'
        add_title('Sales Orders Report', f'Generated on {now}  |  Shiv Furniture MINI ERP')
        headers = ['Order #','Customer','Order Date','Expected Delivery','Status','Subtotal','Tax','Total Amount']
        for c, h in enumerate(headers, 1):
            ws.cell(row=4, column=c, value=h)
        style_header(4, len(headers))
        rows = qry("""SELECT so.order_number, c.company_name, so.order_date,
                      so.expected_delivery_date, so.status, so.subtotal,
                      so.tax_amount, so.total_amount
                      FROM sales_orders so LEFT JOIN customers c ON so.customer_id=c.id
                      ORDER BY so.order_date DESC""")
        for i, r in enumerate(rows, 5):
            ws.cell(row=i, column=1, value=r.get('order_number',''))
            ws.cell(row=i, column=2, value=r.get('company_name',''))
            ws.cell(row=i, column=3, value=str(r.get('order_date','') or ''))
            ws.cell(row=i, column=4, value=str(r.get('expected_delivery_date','') or ''))
            ws.cell(row=i, column=5, value=r.get('status',''))
            ws.cell(row=i, column=6, value=float(r.get('subtotal') or 0))
            ws.cell(row=i, column=7, value=float(r.get('tax_amount') or 0))
            ws.cell(row=i, column=8, value=float(r.get('total_amount') or 0))
            style_data_row(i, len(headers), i%2==0)
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 14
        for col in ['F','G','H']: ws.column_dimensions[col].width = 16
        fname = 'sales_report.xlsx'

    elif rtype == 'purchase':
        ws.title = 'Purchase Report'
        add_title('Purchase Orders Report', f'Generated on {now}  |  Shiv Furniture MINI ERP')
        headers = ['Order #','Vendor','Order Date','Expected Receipt','Status','Subtotal','Tax','Total Amount']
        for c, h in enumerate(headers, 1):
            ws.cell(row=4, column=c, value=h)
        style_header(4, len(headers))
        rows = qry("""SELECT po.order_number, v.company_name, po.order_date,
                      po.expected_receipt_date, po.status, po.subtotal,
                      po.tax_amount, po.total_amount
                      FROM purchase_orders po LEFT JOIN vendors v ON po.vendor_id=v.id
                      ORDER BY po.order_date DESC""")
        for i, r in enumerate(rows, 5):
            ws.cell(row=i, column=1, value=r.get('order_number',''))
            ws.cell(row=i, column=2, value=r.get('company_name',''))
            ws.cell(row=i, column=3, value=str(r.get('order_date','') or ''))
            ws.cell(row=i, column=4, value=str(r.get('expected_receipt_date','') or ''))
            ws.cell(row=i, column=5, value=r.get('status',''))
            ws.cell(row=i, column=6, value=float(r.get('subtotal') or 0))
            ws.cell(row=i, column=7, value=float(r.get('tax_amount') or 0))
            ws.cell(row=i, column=8, value=float(r.get('total_amount') or 0))
            style_data_row(i, len(headers), i%2==0)
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 28
        for col in ['C','D']: ws.column_dimensions[col].width = 16
        ws.column_dimensions['E'].width = 14
        for col in ['F','G','H']: ws.column_dimensions[col].width = 16
        fname = 'purchase_report.xlsx'

    elif rtype == 'inventory':
        ws.title = 'Inventory Report'
        add_title('Inventory Stock Report', f'Generated on {now}  |  Shiv Furniture MINI ERP')
        headers = ['Product Code','Product Name','Category','On Hand Qty','Reserved Qty','Free Qty','Reorder Point','Unit Cost','Inventory Value','Status']
        for c, h in enumerate(headers, 1):
            ws.cell(row=4, column=c, value=h)
        style_header(4, len(headers))
        rows = qry("""SELECT p.product_code, p.product_name, COALESCE(cat.name,'') as category,
                      p.on_hand_qty, p.reserved_qty,
                      GREATEST(0, p.on_hand_qty - p.reserved_qty) as free_qty,
                      p.reorder_point, p.cost_price,
                      (p.on_hand_qty * p.cost_price) as inventory_value,
                      CASE WHEN p.on_hand_qty<=0 THEN 'Out of Stock'
                           WHEN p.on_hand_qty<=p.reorder_point AND p.reorder_point>0 THEN 'Low Stock'
                           ELSE 'OK' END as stock_status
                      FROM products p LEFT JOIN categories cat ON p.category_id=cat.id
                      WHERE p.is_active=1 ORDER BY p.product_name""")
        for i, r in enumerate(rows, 5):
            for c, key in enumerate(['product_code','product_name','category','on_hand_qty',
                                     'reserved_qty','free_qty','reorder_point','cost_price',
                                     'inventory_value','stock_status'], 1):
                val = r.get(key)
                ws.cell(row=i, column=c, value=float(val) if isinstance(val, (int,float)) or
                        (isinstance(val, str) and val.replace('.','',1).lstrip('-').isdigit()) else (val or ''))
            style_data_row(i, len(headers), i%2==0)
            # Color stock status cell
            status_cell = ws.cell(row=i, column=10)
            if status_cell.value == 'Out of Stock':
                status_cell.fill = PatternFill('solid', fgColor='FEE2E2')
                status_cell.font = Font(color='991B1B', bold=True)
            elif status_cell.value == 'Low Stock':
                status_cell.fill = PatternFill('solid', fgColor='FEF3C7')
                status_cell.font = Font(color='92400E', bold=True)
            else:
                status_cell.fill = PatternFill('solid', fgColor='D1FAE5')
                status_cell.font = Font(color='065F46', bold=True)
        col_widths = [14,30,18,12,12,10,12,12,16,14]
        for c, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w
        fname = 'inventory_report.xlsx'

    elif rtype == 'stock-ledger':
        ws.title = 'Stock Ledger'
        add_title('Stock Ledger Report', f'Generated on {now}  |  Shiv Furniture MINI ERP')
        headers = ['Date','Product','Movement Type','Quantity','Balance Qty','Unit Cost','Reference','Notes']
        for c, h in enumerate(headers, 1):
            ws.cell(row=4, column=c, value=h)
        style_header(4, len(headers))
        rows = qry("""SELECT sl.movement_date, p.product_name, sl.movement_type,
                      sl.quantity, sl.balance_qty, sl.unit_cost,
                      COALESCE(sl.reference_number,'') as reference_number,
                      COALESCE(sl.notes,'') as notes
                      FROM stock_ledger sl LEFT JOIN products p ON sl.product_id=p.id
                      ORDER BY sl.movement_date DESC LIMIT 2000""")
        for i, r in enumerate(rows, 5):
            ws.cell(row=i, column=1, value=str(r.get('movement_date','') or ''))
            ws.cell(row=i, column=2, value=r.get('product_name',''))
            ws.cell(row=i, column=3, value=(r.get('movement_type','') or '').replace('_',' ').title())
            ws.cell(row=i, column=4, value=float(r.get('quantity') or 0))
            ws.cell(row=i, column=5, value=float(r.get('balance_qty') or 0))
            ws.cell(row=i, column=6, value=float(r.get('unit_cost') or 0))
            ws.cell(row=i, column=7, value=r.get('reference_number',''))
            ws.cell(row=i, column=8, value=r.get('notes',''))
            style_data_row(i, len(headers), i%2==0)
        col_widths = [18,28,20,12,12,12,18,30]
        for c, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w
        fname = 'stock_ledger.xlsx'

    else:
        return err(f'Unknown report type: {rtype}', 400)

    # Freeze top rows
    ws.freeze_panes = 'A5'

    # Write to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from flask import Response
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={fname}'}
    )

# ── HEALTH ────────────────────────────────────────────────────────────
@app.route('/api/health')
def health():
    return ok({'status':'ok','service':'Mini ERP (Flask)','version':'2.0.0'})

# ── Error Handlers ────────────────────────────────────────────────────
@app.errorhandler(404)
def not_found(e):
    if request.path.startswith('/api/'):
        return err('Not found', 404)
    return send_from_directory(STATIC_DIR, 'index.html')

@app.errorhandler(500)
def server_error(e):
    return err(str(e), 500)

# ── Run ───────────────────────────────────────────────────────────────
if __name__ == '__main__':
    print("=" * 48)
    print("   MINI ERP - Shiv Furniture  (Flask)")
    print("=" * 48)
    print("  App  : http://localhost:8000")
    print("  Login: admin / Admin@123")
    print("=" * 48)
    print("  Press Ctrl+C to stop")
    print()
    app.run(host='0.0.0.0', port=8000, debug=True)
